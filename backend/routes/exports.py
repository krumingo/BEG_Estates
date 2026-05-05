"""Admin Excel/CSV export endpoints."""
import io
from datetime import datetime, timezone
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from auth.dependencies import require_staff
from constants import PROPERTY_STATUS_LABELS, PROPERTY_TYPE_LABELS_BG
from db import get_db

router = APIRouter(prefix="/admin", tags=["admin-export"])


# Per-project floor → kote map (mirrors frontend FLOOR_INFO).
FLOOR_KOTE = {
    -1: "-3.48", 0: "0.00", 1: "+3.40", 2: "+6.29", 3: "+9.18",
    4: "+12.07", 5: "+14.96", 6: "+17.85", 7: "+20.74",
}
FLOOR_LABEL = {
    -1: "СУТЕРЕН", 0: "ПАРТЕР", 1: "1 ЕТАЖ", 2: "2 ЕТАЖ", 3: "3 ЕТАЖ",
    4: "4 ЕТАЖ", 5: "5 ЕТАЖ", 6: "6 ЕТАЖ", 7: "7 ЕТАЖ",
}


def _code_sort_key(code: str):
    """Sort numerically when possible, else alphabetic by Bulgarian collation."""
    try:
        return (0, int(code))
    except (TypeError, ValueError):
        return (1, str(code).lower())


@router.get("/projects/{project_id}/properties/export")
async def export_properties(
    project_id: str,
    format: Literal["xlsx", "csv"] = "xlsx",
    user=Depends(require_staff()),
):
    db = get_db()
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Проектът не е намерен")

    props = await db.properties.find({"project_id": project_id}, {"_id": 0}).to_list(5000)
    props.sort(key=lambda p: (p.get("floor", 0), _code_sort_key(p.get("code") or "")))

    buyer_ids = [p.get("buyer_id") for p in props if p.get("buyer_id")]
    buyers = (
        await db.buyers.find({"id": {"$in": buyer_ids}}, {"_id": 0}).to_list(500)
        if buyer_ids
        else []
    )
    buyer_by_id = {b["id"]: b for b in buyers}

    headers = [
        "Код", "Тип", "Етаж", "Кота", "Стаи",
        "F1 (м²)", "F2 (м²)", "F1+F2 (м²)", "Изложение",
        "Идеални части (%)", "Базова (€)", "Листова (€)",
        "Статус", "Купувач", "Бележки",
    ]

    rows: list[list] = []
    for p in props:
        floor = p.get("floor")
        buyer = buyer_by_id.get(p.get("buyer_id") or "")
        rows.append([
            p.get("code") or "",
            PROPERTY_TYPE_LABELS_BG.get(p.get("property_type") or "", p.get("property_type") or ""),
            FLOOR_LABEL.get(floor, str(floor) if floor is not None else ""),
            FLOOR_KOTE.get(floor, ""),
            p.get("rooms") if p.get("property_type") == "apartment" else None,
            p.get("raw_area"),
            p.get("ideal_parts_area"),
            p.get("area_total"),
            p.get("exposure") or "",
            p.get("ideal_parts_area"),
            p.get("base_price"),
            p.get("list_price"),
            PROPERTY_STATUS_LABELS.get(p.get("status") or "available", p.get("status") or ""),
            buyer.get("name") if buyer else "",
            p.get("admin_notes") or "",
        ])

    slug = project.get("slug") or project.get("id") or "inventory"
    date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"{slug}-inventar-{date}"

    if format == "csv":
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Инвентар"

    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="1E293B")  # slate-800
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    numeric_cols = {5, 6, 7, 8, 10, 11, 12}  # rooms, F1, F2, F1+F2, ideal, base, list
    money_cols = {11, 12}
    area_cols = {6, 7, 8}

    for r in rows:
        ws.append(r)

    for r_i in range(2, ws.max_row + 1):
        for c_i in range(1, len(headers) + 1):
            cell = ws.cell(row=r_i, column=c_i)
            if c_i in money_cols:
                if cell.value is not None and cell.value != "":
                    cell.number_format = "#,##0 €"
                cell.alignment = right
            elif c_i in area_cols:
                if cell.value is not None and cell.value != "":
                    cell.number_format = "0.00"
                cell.alignment = right
            elif c_i in numeric_cols:
                cell.alignment = right
            else:
                cell.alignment = left

    # Column widths (auto-fit best-effort)
    widths = [12, 22, 12, 10, 6, 10, 10, 10, 16, 10, 12, 12, 16, 24, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
